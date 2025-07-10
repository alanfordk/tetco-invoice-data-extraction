#!/usr/bin/env python3
import os
import re
import sys
import tempfile
import subprocess
import tkinter as tk
from tkinter import filedialog

import pandas as pd
import fitz  # PyMuPDF
from openpyxl import load_workbook

# ─── Ensure we can import your readers/ folder as a package ─────────────────
root_dir = os.path.dirname(__file__)
sys.path.insert(0, os.path.join(root_dir, "readers"))
from invoices.pdf_reader import extract_invoice_data
from invoices.wpd_reader import extract_invoice_data as extract_wpd_data

# ─── All clients & their candidate locations ─────────────────────────────────
ALL_LOCATIONS = {
    'JR Simplot':       ['Pocatello, ID', 'Caldwell, ID', 'Aberdeen, ID', 'Overton, NV', 'Afton, WY'],
    'Intrepid Potash':  ['Moab, UT', 'Wendover, NV'],
    'Lamb Weston Inc':  ['Twin Falls, ID', 'American Falls, ID', 'Qincy, WA'],
}

def find_location_in_text(text, location_list):
    for loc in location_list:
        city = loc.split(',')[0]
        if re.search(rf'\b{re.escape(city)}\b', text, re.IGNORECASE):
            return loc
    return None

def extract_raw_text(path):
    doc = fitz.open(path)
    return "\n".join(page.get_text() or "" for page in doc)

def extract_text_from_wpd(path):
    with tempfile.TemporaryDirectory() as tmpdir:
        subprocess.run([
            "soffice", "--headless", "--convert-to", "pdf",
            "--outdir", tmpdir, path
        ], check=True)
        name, _ = os.path.splitext(os.path.basename(path))
        return extract_raw_text(os.path.join(tmpdir, name + ".pdf"))

def main():
    # 1) Pick folder with invoices
    root = tk.Tk(); root.withdraw()
    folder = filedialog.askdirectory(title="Select folder with invoice files")
    if not folder:
        print("No folder selected. Exiting."); sys.exit(1)

    # 2) Hard-coded Excel path
    wb_path = r"C:\Users\akitc\OneDrive\Desktop\Data project\Tetco_invoices.xlsx"

    # 3) Load your full invoice sheet into pandas (for filtering only)
    df = pd.read_excel(wb_path)

    # 4) Choose which client to process
    clients = list(ALL_LOCATIONS.keys())
    print("Which client do you want to search for?")
    for i, name in enumerate(clients, start=1):
        print(f"  {i}. {name}")
    choice = input("Enter number: ").strip()
    try:
        sel = clients[int(choice)-1]
    except Exception:
        print("Invalid choice. Exiting."); sys.exit(1)
    loc_list = ALL_LOCATIONS[sel]
    df_sub = df[df['client_name'] == sel].copy()

    # 5) Build regex and results dict
    pattern = re.compile(r'INVOICE\D*(\d+)', re.IGNORECASE)
    invoice_to_loc = {}

    # 6) Walk and infer
    for root_dir, _, files in os.walk(folder):
        for fn in files:
            ext = os.path.splitext(fn)[1].lower()
            if ext not in ('.pdf', '.wpd', '.wps'): continue

            m = pattern.search(fn)
            if not m: continue
            inv = m.group(1).lstrip('0')
            if not any(str(x).lstrip('0') == inv for x in df_sub['invoice_number']):
                continue

            path = os.path.join(root_dir, fn)

            # a) filename
            loc = find_location_in_text(fn, loc_list)

            # b) parsed descriptions
            if not loc:
                try:
                    recs = (extract_invoice_data(path)
                            if ext=='.pdf'
                            else extract_wpd_data(path))
                    blob = " ".join(r.get('project_description','') for r in recs)
                except Exception:
                    blob = ""
                loc = find_location_in_text(blob, loc_list)

            # c) full-text fallback
            if not loc:
                text = (extract_raw_text(path)
                        if ext=='.pdf'
                        else extract_text_from_wpd(path))
                loc = find_location_in_text(text, loc_list)

            if loc:
                invoice_to_loc[inv] = loc

    # 7) Report
    print("\nInferred locations:")
    for inv, l in invoice_to_loc.items():
        print(f"  • {sel} Invoice {inv}: {l}")
    if not invoice_to_loc:
        print("  (none found)")

    # 8) Update workbook in place
    wb = load_workbook(wb_path)
    ws = wb.active  # or wb['SheetName'] if you need a specific sheet

    # Map headers to column indices
    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    col_map = {cell.value: idx for idx, cell in enumerate(header_row, start=1)}
    inv_col = col_map['invoice_number']
    loc_col = col_map['location']

    for row in ws.iter_rows(min_row=2):
        inv_cell = row[inv_col-1]
        key = str(inv_cell.value).lstrip('0')
        if key in invoice_to_loc:
            row[loc_col-1].value = invoice_to_loc[key]

    wb.save(wb_path)
    print(f"\nUpdated workbook in place → {wb_path}")

if __name__ == "__main__":
    main()
