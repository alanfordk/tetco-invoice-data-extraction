import os
import sys
import tkinter as tk
from tkinter import filedialog, simpledialog

import openpyxl

from invoices.pdf_reader import extract_invoice_data as extract_pdf
from invoices.wpd_reader import extract_invoice_data as extract_wpd

# Path to your fixed Excel workbook
WORKBOOK_PATH = r"C:\Users\akitc\OneDrive\Desktop\Data project\Tetco_invoices.xlsx"

# Map file extensions to reader functions
READERS = {
    '.pdf': extract_pdf,
    '.wpd': extract_wpd,
}

def process_folder(folder_path):
    """Recursively parse all supported invoices in folder_path."""
    rows = []
    for dirpath, _, filenames in os.walk(folder_path):
        for fname in filenames:
            ext = os.path.splitext(fname)[1].lower()
            reader = READERS.get(ext)
            if not reader:
                continue
            full = os.path.join(dirpath, fname)
            try:
                data = reader(full)
                items = data if isinstance(data, list) else [data]
                for rec in items:
                    rec['source_file'] = os.path.relpath(full, folder_path)
                    rows.append(rec)
            except Exception as e:
                print(f"Error processing {full}: {e}")
    return rows

def append_to_workbook(rows):
    """Append rows into the fixed Excel workbook with formatting."""
    cols = [
        "client_name",
        "project_description",
        "test_date_start",
        "test_date_end",
        "mobilization_count",
        "invoice_amount",
        "invoice_date",
        "invoice_number",
        "location",
    ]
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    ws = wb.active
    start_row = ws.max_row + 1

    for rec in rows:
        for col_idx, key in enumerate(cols, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=rec.get(key))

            # Wrap only the description column
            if key == "project_description":
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            else:
                cell.alignment = openpyxl.styles.Alignment(
                    horizontal="center", vertical="center"
                )

            if key == "invoice_amount":
                cell.number_format = '"$"#,##0.00'
            elif key in ("test_date_start", "test_date_end", "invoice_date"):
                cell.number_format = "MM/DD/YYYY"

        start_row += 1

    wb.save(WORKBOOK_PATH)
    print(f"\nAppended {len(rows)} records to:\n  {WORKBOOK_PATH}")

def cli_mode(argv):
    # 1) Prompt for client info
    client_name = input("Enter client name: ").strip()
    location    = input("Enter location: ").strip()

    # 2) Determine folder and process
    folder = argv[1] if len(argv) > 1 else 'invoices'
    print(f"\nProcessing invoices in folder:\n  {folder}\n")
    rows = process_folder(folder)
    if not rows:
        print("No records found. Exiting.")
        return

    # 3) Stamp every record with client info
    for rec in rows:
        rec['client_name'] = client_name
        rec['location']    = location

    # 4) Preview first 20
    preview = rows[:20]
    print(f"=== Previewing {len(preview)} of {len(rows)} records ===")
    for rec in preview:
        print(", ".join(f"{k}={v}" for k, v in rec.items()))

    # 5) Append to workbook
    append_to_workbook(rows)

def gui_mode():
    root = tk.Tk()
    root.withdraw()

    # 1) Pick invoices folder
    folder = filedialog.askdirectory(title="Select Invoices Folder")
    if not folder:
        print("No folder selected. Exiting.")
        return

    # 2) Ask for client info via dialog
    client_name = simpledialog.askstring("Client Name", "Enter client name:")
    if not client_name:
        print("No client name entered. Exiting.")
        return
    location = simpledialog.askstring("Location", "Enter location:")
    if not location:
        print("No location entered. Exiting.")
        return

    # 3) Process folder
    rows = process_folder(folder)
    if not rows:
        print("No invoices found in that folder. Exiting.")
        return

    # 4) Stamp client info
    for rec in rows:
        rec['client_name'] = client_name
        rec['location']    = location

    # 5) Preview
    preview = rows[:20]
    print(f"=== Previewing {len(preview)} of {len(rows)} records ===")
    for rec in preview:
        print(", ".join(f"{k}={v}" for k, v in rec.items()))

    # 6) Append to workbook
    append_to_workbook(rows)

if __name__ == "__main__":
    # CLI if any args, otherwise GUI
    if len(sys.argv) > 1:
        cli_mode(sys.argv)
    else:
        gui_mode()
